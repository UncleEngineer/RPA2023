from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time

import pyautogui as pg
from openpyxl import load_workbook

loadfile = 'stockprice.xlsx'
myworkbook = load_workbook(loadfile)
mysheet = myworkbook.active
lastrow = len(mysheet['A']) + 1

stocklist = []
pricelist = []

for i in range(2, lastrow):
    stocklist.append(mysheet.cell(row=i, column=1).value)

path = r'C:\Users\Uncle Engineer\Desktop\RPA 2023\chromedriver.exe'
service = Service(path)

option = webdriver.ChromeOptions()
option.add_argument('--start-maximized')

driver = webdriver.Chrome(service=service, options=option)

for s in stocklist:
    driver.get("https://www.google.com")

    stock = driver.find_element(By.NAME, 'q')
    stock.send_keys(f'ราคาหุ้น {s}')
    stock.send_keys(Keys.RETURN)

    xpath = '//*[@id="knowledge-finance-wholepage__entity-summary"]/div[3]/g-card-section/div/g-card-section/div[2]/div[1]/span[1]/span/span[1]'
    price = driver.find_element(By.XPATH, xpath).text
    pricelist.append(float(price))

    for r, p in enumerate(pricelist, 2):
        writedata = mysheet.cell(row=r, column=2)
        writedata.value = p

    img_search = pg.screenshot(region=(208, 311, 697, 582))
    img_search.save(f'{s}.png')

    time.sleep(5)
    print(s, p)

driver.quit()
myworkbook.save(loadfile)